#from flask import Flask,render_template
from flask_mysqldb import MySQL
from flask import Flask, render_template, flash, request, url_for, redirect, session
from flask_wtf import FlaskForm, Form
from wtforms import BooleanField, TextField, PasswordField, validators, SubmitField, StringField
from MySQLdb import escape_string as thwart
import gc
from flask import g
from datetime import datetime
from datetime import date
#from werkzeug import secure_filename
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from werkzeug.utils import secure_filename
from os import getcwd
from pathvalidate import sanitize_filename
from flask import send_file

app = Flask(__name__)

app.config['TESTING'] = True

app.config['MYSQL_HOST']='localhost'
app.config['MYSQL_USER']='root'
app.config['MYSQL_PASSWORD']=''
app.config["MYSQL_DB"]='ffs'
app.config['SECRET_KEY']='SQLIDP'
mysql = MySQL(app)

#class RegistrationForm():
    #id = TextField('Employee id', [validators.Length(min=4, max=20)])
    #psw = PasswordField('Password', [validators.Length(min=6, max=50)])
    #submit = SubmitField('Register')
    #password = PasswordField('psw-repeat', [
    #    validators.Required(),
    #    validators.EqualTo('psw', message='Passwords must match')
    #])
    #confirm = PasswordField('Repeat Password')
    #accept_tos = BooleanField('I accept the Terms of Service and Privacy Notice (updated Jan 22, 2015)', [validators.Required()])

@app.route("/")
def home():
    return render_template('duplicate.html')
'''
@app.route('/login',methods=['GET','POST'])
def login():
    if request.method == 'POST':
        id = request.form['id']
        psw = request.form['password']
        c = mysql.connection.cursor()
        c.excute(select * from user1 where id=%s,(id,))
        x=c.fetchall()
        count = 0
        for k in x:
            count=count+1
        if count > 0:

        return render_template('login.html')
'''

@app.route("/about")
def about():
    return render_template("about.html")


@app.route('/index')
def index():
    if 'username' in session:
      username = session['username']
      return render_template('index.html',username=username)
'''
      return 'Logged in as ' + username + '<br>' + \
         "<b><a href = '/logout'>click here to log out</a></b>"
    return "You are not logged in <br><a href = '/login'></b>" + \
      "click here to log in</b></a>"
    return render_template('index.html')
'''
@app.route('/logout')
def logout():
   # remove the username from the session if it is there
   session.pop('username', None)
   #headers.add('Cache-Control', 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0')
   #response.headers.add('Cache-Control', 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0')
   return redirect(url_for('home'))

@app.after_request
def after_request(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response

@app.after_request
def after_this_request(func):
    if not hasattr(g, 'call_after_request'):
        g.call_after_request = []
    g.call_after_request.append(func)
    return func


@app.after_request
def per_request_callbacks(response):
    for func in getattr(g, 'call_after_request', ()):
        response = func(response)
    return response

@app.route('/adminindex')
def adminindex():
    return render_template('admin.html')

@app.route('/hod',methods=['GET','POST'])
def hod():
    error = None
    if request.method == 'POST':
        username_form  = request.form['id']
        password_form  = request.form['password']
        c = mysql.connection.cursor()
        c.execute('''SELECT * FROM user2 WHERE id = %s and psw= %s''', (username_form, password_form,)) # CHECKS IF USERNAME EXSIST
        x=c.fetchall()
        count = 0
        for k in x:
            count=count+1
        if count > 0:
            username = username_form
            session['username'] = username_form
            return render_template('adminhod.html',username=username)
        else:
            error = "Invalid Credential"
            #flash('Invalid Credentials ... !!!')
            return render_template('hod.html', error=error)
    return render_template('hod.html')

@app.route('/loginspecial', methods=['GET', 'POST'])
def loginspecial():
    error = None
    if request.method == 'POST':
        username_form  = request.form['id']
        password_form  = request.form['password']
        c = mysql.connection.cursor()
        c.execute('''SELECT * FROM user2 WHERE id = %s and psw= %s''', (username_form, password_form,)) # CHECKS IF USERNAME EXSIST
        x=c.fetchall()
        count = 0
        for k in x:
            count=count+1
        if count > 0:
            username = username_form
            session['username'] = username_form
            return render_template('admin.html',username=username)
        else:
            error = "Invalid Credential"
            #flash('Invalid Credentials ... !!!')
            return render_template('loginspecial.html', error=error)
    else:
        return render_template('loginspecial.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username_form  = request.form['id']
        password_form  = request.form['password']
        c = mysql.connection.cursor()
        c.execute('''SELECT * FROM user1 WHERE id = %s and psw= %s''', (username_form, password_form,)) # CHECKS IF USERNAME EXSIST
        x=c.fetchall()
        count = 0
        for k in x:
            count=count+1
        if count > 0:
            session['username'] = username_form
            return redirect(url_for('index'))
        else:
            error = "Invalid Credential"
            return render_template('login.html', error=error)
    else:
        return render_template('login.html')

@app.route('/viewprofile',methods=['GET','POST'])
def viewprofile():
    username = session.get('username',None)
    l = []
    x = mysql.connection.cursor()
    x.execute('select * from profiles where id=%s',[username])
    data = x.fetchall()
    k = len(data)
    if k != 0:
        l = list(data)
    else:
        l = [[username,'Not Entered','Not Entered','Not Entered','Not Entered','Not Entered','Not Entered','Not Entered','Not Entered','Not Entered','Not Entered','Not Entered']]
    return render_template('viewprofile.html',l=l,username=username)


@app.route('/addhod', methods=['GET','POST'])
def addhod():
    x = mysql.connection.cursor()
    x.execute('select id from user2')
    y = x.fetchall()
    x.close()
    gc.collect()
    l = []
    count = 0
    for z in y:
        str = ''
        str = ''.join(z)
        l.append(str)
        count = count +1
    username = session.get('username',None)
    if request.method == 'POST':
        hodname = request.form['id']
        x = mysql.connection.cursor()
        x.execute('select * from user2 where id = %s',(hodname,))
        y = x.fetchall()
        count = 0
        for k in y:
            count = count + 1
        if count > 0:
            flash('HoD account already exists ... !!!!')
            return render_template('addhod.html',username=username,l=l)
        else:
            x.execute('insert into user2(id,psw) values (%s,%s)',(hodname,hodname))
            x.connection.commit()
            x.close()
            flash('HoD account successfully created ... !!!')
            gc.collect()
            x = mysql.connection.cursor()
            x.execute('select id from user2')
            y = x.fetchall()
            x.close()
            gc.collect()
            l = []
            count = 0
            for z in y:
                str = ''
                str = ''.join(z)
                l.append(str)
                count = count +1
            return render_template('addhod.html',username=username,l=l)
    else:
        return render_template('addhod.html',username=username,l=l)

@app.route('/deletehod',methods=['GET','POST'])
def deletehod():
    username = session.get('username',None)
    x = mysql.connection.cursor()
    x.execute('select id from user2')
    y = x.fetchall()
    x.close()
    gc.collect()
    l = []
    count = 0
    for z in y:
        str = ''
        str = ''.join(z)
        l.append(str)
        count = count +1
    if request.method == 'POST':
        delname = request.form['id']
        c=0
        if delname == 'admin':
            flash('You cannot delete admin account')
            return render_template('deletehod.html',username=username,l=l,count=count)
        elif delname in l:
            x = mysql.connection.cursor()
            x.execute('delete from user2 where id = %s',(delname,))
            x.connection.commit()
            x.close()
            flash("HoD account successfully deleted ... !!!")
            gc.collect()
            x = mysql.connection.cursor()
            x.execute('select id from user2')
            y = x.fetchall()
            x.close()
            gc.collect()
            l = []
            count = 0
            for z in y:
                str = ''
                str = ''.join(z)
                l.append(str)
                count = count +1
            return render_template('deletehod.html',username=username,l=l,count=count)
        else:
            flash('HoD account is not available .... Enter valid HoD account')
            return render_template('deletehod.html',username=username,l=l,count=count)
    else:
        return render_template('deletehod.html',username=username,l=l,count=count)




@app.route('/profile',methods=['GET','POST'])
def profile():
    username=session.get('username',None)
    l = []
    x = mysql.connection.cursor()
    x.execute('select * from profiles where id=%s',[username])
    data = x.fetchall()
    k = len(data)
    x.close()
    gc.collect()
    if k != 0:
        l = list(data)
    else:
        l = [[username,'Not Entered','Not Entered','Not Entered','Not Entered','Not Entered','Not Entered','Not Entered','Not Entered','Not Entered','Not Entered','Not Entered']]
    x = mysql.connection.cursor()
    x.execute('''select * from profiles where id = %s ''',(username,))
    y = x.fetchall()
    x.close()
    gc.collect()
    count = 0
    for k in y:
        count = count +1
    if  count > 0:
        flash('User Profile already update')
        return render_template('updateprofile.html',username=username,l=l)
    else:
        if request.method == 'POST':
            x = mysql.connection.cursor()
            x.execute('''select * from profiles where id = %s ''',(username,))
            y = x.fetchall()
            count = 0
            for k in y:
                count =count +1
            if count > 0:
                flash("User profile already exists....!!!!")
                return redirect(url_for('index'))
            else:
                name = request.form['name']
                designation=request.form['designation']
                dob1 = request.form['dob']
                dob = datetime.strptime(dob1,"%Y-%m-%d")
                doj1 = request.form['doj']
                doj = datetime.strptime(doj1,"%Y-%m-%d")
                phone = request.form['phone']
                email = request.form['email']
                btech = request.form['btech']
        #btech = datetime.strptime(b,"%d/%m/%Y %H:%M")
        #btech = float(b)
                mtech = request.form['mtech']
        #mtech = datetime.strptime(m,"%d/%m/%Y %H:%M")
        #mtech=float(m)
                other = request.form['exp_other']
                adhar = request.form['adhar']
                pan = request.form['pan']
                c = mysql.connection.cursor()
                c.execute('insert into profiles(id,name,designation,dob,doj,phone,email,btech,mtech,exp_other,adhar,pan) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',(username,name,designation,(dob.strftime('%Y-%m-%d')),(doj.strftime('%Y-%m-%d')),phone,email,btech,mtech,other,adhar,pan))
                c.connection.commit()
                c.close()
                gc.collect()
                flash("Profile successfully created......")
                return redirect(url_for('index'))
    return render_template('profile.html',username=username)

@app.route('/updateprofile', methods=['GET','POST'])
def updateprofile():
    username = session.get('username',None)
    l = [[username,'Not Entered','Not Entered','Not Entered','Not Entered','Not Entered','Not Entered','Not Entered','Not Entered','Not Entered','Not Entered','Not Entered']]
    x = mysql.connection.cursor()
    x.execute('select * from profiles where id=%s',[username])
    data = x.fetchall()
    k = len(data)
    x.close()
    gc.collect()
    l = list(data)
    if request.method == 'POST':
        select = request.form['sel']
        value = request.form['val']
        value1 = request.form['designation']
        if (select == 'none') or (select !='designation' and value == '') or (select =='designation' and value1 == ''):
            flash("please enter a valid value");
            return render_template('updateprofile.html',username=username,l=l)
        x = mysql.connection.cursor()
        if select == 'name':
            x.execute('update profiles set name=%s where id=%s',(value,username))
        elif select == 'designation':
            value = request.form['designation']
            x.execute('update profiles set designation=%s where id=%s',(value,username))
        elif select =='dob':
            x.execute('update profiles set dob=%s where id=%s',(value,username))
        elif select == 'doj':
            x.execute('update profiles set doj=%s where id=%s',(value,username))
        elif select == 'phone':
            x.execute('update profiles set phone=%s where id=%s',(value,username))
        elif select == 'email':
            x.execute('update profiles set email=%s where id=%s',(value,username))
        elif select == 'btech':
            x.execute('update profiles set btech=%s where id=%s',(value,username))
        elif select == 'mtech':
            x.execute('update profiles set mtech=%s where id=%s',(value,username))
        elif select =='exp_other':
            x.execute('update profiles set exp_other=%s where id=%s',(value,username))
        elif select == 'adhar':
            x.execute('update profiles set adhar=%s where id=%s',(value,username))
        elif select == 'pan':
            x.execute('update profiles set pan=%s where id=%s',(value,username))
        x.connection.commit()
        x.close()
        gc.collect()
        x = mysql.connection.cursor()
        x.execute('select * from profiles where id=%s',[username])
        data = x.fetchall()
        k = len(data)
        x.close()
        gc.collect()
        l = list(data)
        flash('Field value updated successfully... !!!')
        return render_template('updateprofile.html',l=l,username=username)
    else:
        return render_template('updateprofile.html',l=l,username=username)
    return render_template('updateprofile.html',l=l,username=username)


@app.route('/register',methods=['GET','POST'])
def register():
        #form = RegistrationForm()
        if request.method == 'POST':
            eid = request.form['id']
            psw = request.form['psw']
            error=None
            #id  = request.form.get()
            #psw= from.psw.data
            #password = request.form.get()#sha256_crypt.encrypt((str(form.psw.data)))
            #c = mysql.connection.cursor()
            c= mysql.connection.cursor()
            c.execute('''SELECT * FROM user1 WHERE id = %s''',(eid,))
            x=c.fetchall()
            count = 0
            for k in x:
                count=count+1
            if count > 0:
                error="USER EXISTS !!!!!"
                flash("That username is already exists......!!!!")
                return render_template('register.html')
            else:
                c.execute('''INSERT INTO user1 (id, psw) VALUES (%s, %s)''',(eid, psw))
                c.connection.commit()
                c.close()
                #conn.close()
                gc.collect()
                session['logged_in'] = True
                session['id'] = eid
                flash('User successfully created.....!')
                flash("Thanks for registering!")
                return redirect(url_for('home'))
        else:
            error = "User already exists....!!!!"
            return render_template("register.html",form=Form)

    #except Exception as e:
    #    return(str(e))
            #return 'hello'
            #return render_template('register.html',form=form)


def convertTuple(tup):
    str =  ''.join(tup)
    return str

class MyForm(Form):
    val = StringField('newpass')
    val1 = StringField('newpass1')
    button = SubmitField('submitinner')


@app.route('/changepassword',methods=['GET','POST'])
def changepassword():
    password = None
    msg = None
    script = False
    username = session.get('username')
    if request.method == 'POST':
        password = request.form['oldpass']
        x = mysql.connection.cursor()
        x.execute('select psw from user1 where id=%s',(username,))
        data = x.fetchall()
        if password in data[0]:
            return render_template('changepassword1.html',username=username)
        elif password == '':
            flash('Password field is empty... Please enter a valid password ...!!!!')
            return render_template('changepassword.html',username=username)
        else:
            flash('Mismatch of passwords entered, reenter the passwords !!!!.....')
            return render_template('changepassword.html',username=username)
    else:
        return render_template('changepassword.html',username=username)

@app.route('/changepassword1',methods=['GET','POST'])
def changepassword1():
    username = session.get('username')
    if request.method == 'POST':
        password = request.form['newpass']
        password1 = request.form['newpass1']
        if (password == password1 and password != ''):
            x = mysql.connection.cursor();
            x.execute('update user1 set psw = %s where id = %s',(password,username))
            x.connection.commit()
            x.close()
            gc.collect()
            flash('Password updated successfully ...... !!!!')
            return render_template('changepassword1.html',username=username)
        elif password =='':
            flash('Password field is empty... Please enter a valid password ...!!!!')
            return render_template('changepassword1.html',username=username)
        else:
            flash('Mismatch of passwords entered, reenter the passwords !!!!.....')
            return render_template('changepassword1.html',username=username)
    else:
        flash('Something went wrong.... Connection error... refresh the page !!!!!')
        return render_template('changepassword1.html',username=username)

class form1(FlaskForm):
    s1 = SubmitField('submit1')

class form2(FlaskForm):
    s2 = SubmitField('submit2')



@app.route('/addfaculty',methods=['GET','POST'])
def addfaculty():
    f1 = form1()
    f2 = form2()
    data = []
    available = []
    added = []
    script = False
    username = session.get('username',None)
    if request.method == 'POST':
        if request.form['sub'] == 'Upload':
            filepath = "load.xlsx"
            wb = Workbook()
            wb = load_workbook(filepath)
            sheet = wb.active
            r = sheet.max_row
            c = sheet.max_column
            for i in range(1, r+1):
                for j in range(1, c+1):
                    a = sheet.cell(row=i,column=j).value
                    data.append(a)
            for k in data:
                x = mysql.connection.cursor()
                x.execute('''select * from user1 where id = %s''',(k,))
                y = x.fetchall()
                count = 0
                for i in y:
                    count = count + 1
                x.close()
                gc.collect()
                if count>0:
                    available.append(k)
                else:
                    x = mysql.connection.cursor()
                    x.execute('insert into user1(id,psw) values (%s,%s)',(k,k))
                    x.connection.commit()
                    x.close()
                    added.append(k)
            script = True
            return render_template('addfaculty.html',username=username,script=script,available=available,added=added)
        elif request.form['sub'] == 'Click':
            id = request.form['id']
            if id == '':
                flash('ID must not be nul... Enter a valid ID')
                return render_template('addfaculty.html',username=username,data=data)
            else:
                x = mysql.connection.cursor()
                x.execute('select * from user1 where id=%s',(id,))
                y = x.fetchall()
                x.close()
                gc.collect()
                l = []
                count = 0
                for i in y:
                    count = count  + 1
                if count > 0:
                    flash('Faculty already exists ......')
                    return render_template('addfaculty.html',username=username,data=data)
                else:
                    x = mysql.connection.cursor()
                    x.execute('''INSERT INTO user1 (id, psw) VALUES (%s, %s)''',(id, id))
                    x.connection.commit()
                    x.close()
                    flash('Faculty added successfully...')
                    return render_template('addfaculty.html',username=username,data=data)
        else:
            flash('Not in any method.....!!!')
            return render_template('addfaculty.html',username=username,data=data)

    else:
        return render_template('addfaculty.html',username=username,data=data)



@app.route('/viewfaculty',methods=['GET','POST'])
def viewfaculty():
    username = session.get('username',None)
    l = []
    cse  = []
    it  = []
    ece = []
    eee = []
    mba = []
    mca = []
    ce = []
    me = []
    fed = []
    ae = []
    others = []
    script = False
    x = mysql.connection.cursor()
    x.execute('select id from user1')
    y = x.fetchall()
    count = 0
    for i in y:
        l.append(i)
        count = count + 1
    if count > 0:
        script = True
        l.sort()
        for i in l:
            j=''
            j =''.join(i)
            if j.startswith('cse',0,3):
                cse.append(j)
            elif j.startswith('ece',0,3):
                ece.append(j)
            elif j.startswith('it',0,2):
                it.append(j)
            elif j.startswith('eee',0,3):
                eee.append(j)
            elif j.startswith('fed',0,3):
                fed.append(j)
            elif j.startswith('mba',0,3):
                mba.append(j)
            elif j.startswith('mca',0,3):
                mca.append(j)
            elif j.startswith('ce',0,2):
                ce.append(j)
            elif j.startswith('me',0,2):
                me.append(j)
            elif j.startswith('ae',0,2):
                ae.append(j)
            else:
                others.append(j)
    else:
        return render_template('viewfaculty.html',username=username,script=script)
    if request.method == 'POST':
        id = request.form['id']
        if id == '':
            flash('Enter a valid ID......')
            return render_template('viewfaculty.html',username=username,script=script,cse=cse,it=it,ece=ece,eee=eee,fed=fed,mba=mba,mca=mca,ce=ce,me=me,ae=ae,others=others)
        x = mysql.connection.cursor()
        x.execute('select * from user1 where id=%s',(id,))
        y = x.fetchall()
        x.close()
        gc.collect()
        count = 0
        for i in y:
            count = count + 1
        if count > 0:
            x = mysql.connection.cursor()
            x.execute('delete from user1 where id = %s',(id,))
            x.connection.commit()
            x.close()
            flash('Deleted successfullt..')
            gc.collect()
            return render_template('viewfaculty.html',username=username,script=script,cse=cse,it=it,ece=ece,eee=eee,fed=fed,mba=mba,mca=mca,ce=ce,me=me,ae=ae,others=others)
        else:
            flash('Faculty account for entered id is not available...')
            return render_template('viewfaculty.html',username=username,script=script,cse=cse,it=it,ece=ece,eee=eee,fed=fed,mba=mba,mca=mca,ce=ce,me=me,ae=ae,others=others)
    return render_template('viewfaculty.html',username=username,script=script,cse=cse,it=it,ece=ece,eee=eee,fed=fed,mba=mba,mca=mca,ce=ce,me=me,ae=ae,others=others)


@app.route('/cas')
def cas():
    username = session.get('username')
    return render_template('cas.html',username=username)

@app.route('/mycas',methods=['GET','POST'])
def mycas():
    username = session.get('username',None)
    if request.method == 'POST':
        fromdate = request.form['fromdate']
        todate = request.form['todate']
        if fromdate > todate:
            flash('From-Date must be less than To-Date')
            return render_template('mycas.html',username=username)
        else:
            sum1,sum2,sum3=0,0,0
            x = mysql.connection.cursor()
            x.execute("select * from conference where id=%s and cdate>=%s and cdate<=%s",(username,fromdate,todate,))
            y = x.fetchall()
            clen = len(y)
            if(clen==0):
                sum1=0
            else:
                x.execute("select SUM(score) from conference where id=%s and cdate>=%s and cdate<=%s",(username,fromdate,todate,))
                sum1 = int(x.fetchall()[0][0]) 
            x.close()
            gc.collect()
            a = mysql.connection.cursor()
            a.execute("select * from journal where id=%s and cdate>=%s and cdate<=%s",(username,fromdate,todate,))
            b = a.fetchall()
            jlen = len(b)
            if(jlen==0):
                sum2=0
            else:
                a.execute("select SUM(score) from journal where id=%s and cdate>=%s and cdate<=%s",(username,fromdate,todate,))
                sum2 = int(a.fetchall()[0][0]) 
            a.close()
            gc.collect()
            k = mysql.connection.cursor()
            k.execute("select * from fdp where id=%s and fromdate>=%s and fromdate<=%s",(username,fromdate,todate,))
            f = k.fetchall()
            flen = len(f)
            if(flen==0):
                sum3=0
            else:
                k.execute("select SUM(score) from fdp where id=%s and fromdate>=%s and fromdate<=%s",(username,fromdate,todate,))
                sum3 = int(k.fetchall()[0][0]) 
            k.close()
            gc.collect()
            totsum = sum1+sum2+sum3
            return render_template('viewcas.html',username=username,y=y,b=b,clen=clen,jlen=jlen,f=f,flen=flen,sum1=sum1,sum2=sum2,sum3=sum3,totsum=totsum)
        return render_template('mycas.html',username=username)
    return render_template('mycas.html',username=username)

@app.route('/casscore',methods=['GET','POST'])
def casscore():
    username = session.get('username', None)
    x = mysql.connection.cursor()
    flag = x.execute('select * from casmarks')
    if flag == True:
        x.execute("select * from casmarks ")
        y=x.fetchall()
        clen=len(y)
        x.close()
        gc.collect()
        flash('CAS Scores already updated ... to modify scores click on CAS Scores Reset')
        return render_template('existingcas.html',username=username,y=y,clen=clen)
    else:
        if request.method == 'POST':
            ugc = request.form['indexugc']
            scopus = request.form['indexscopus']
            sci = request.form['indexsci']
            oth = request.form['indexother']
            iit = request.form['pointiit']
            iiit = request.form['pointiiit']
            nit = request.form['pointnit']
            univ = request.form['pointuniv']
            duniv = request.form['pointduniv']
            puniv = request.form['pointpuniv']
            clg = request.form['pointother']
            x = mysql.connection.cursor()
            flag = x.execute('select * from casmarks')
            x.close()
            gc.collect()
            if flag == False:
                x = mysql.connection.cursor()
                s = 'insert into casmarks(ugc,scopus,sci,others,iit,iiit,nit,univ,duniv,puniv,college) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
                a = (ugc,scopus,sci,oth,iit,iiit,nit,univ,duniv,puniv,clg,)
                x.execute(s,a)
                x.connection.commit()
                x.close()
                gc.collect()
                flash('CAS Scores stored successfully ... !!!')
                return render_template('cas.html',username=username)
    return render_template('casscore.html',username=username)


@app.route('/casformset')
def casformset():
    username = session.get('username',None)
    l = []
    cse  = []
    it  = []
    ece = []
    eee = []
    mba = []
    mca = []
    ce = []
    me = []
    fed = []
    ae = []
    others = []
    script = False
    x = mysql.connection.cursor()
    x.execute('select id from user1')
    y = x.fetchall()
    count = 0
    for i in y:
        l.append(i)
        count = count + 1
    if count > 0:
        script = True
        l.sort()
        for i in l:
            j=''
            j =''.join(i)
            if j.startswith('cse',0,3):
                cse.append(j)
            elif j.startswith('ece',0,3):
                ece.append(j)
            elif j.startswith('it',0,2):
                it.append(j)
            elif j.startswith('eee',0,3):
                eee.append(j)
            elif j.startswith('fed',0,3):
                fed.append(j)
            elif j.startswith('mba',0,3):
                mba.append(j)
            elif j.startswith('mca',0,3):
                mca.append(j)
            elif j.startswith('ce',0,2):
                ce.append(j)
            elif j.startswith('me',0,2):
                me.append(j)
            elif j.startswith('ae',0,2):
                ae.append(j)
            else:
                others.append(j)
    else:
        return render_template('viewfaculty.html',username=username,script=script)
    if request.method == 'POST':
        id = request.form['id']
        if id == '':
            flash('Enter a valid ID......')
            return render_template('viewfaculty.html',username=username,script=script,cse=cse,it=it,ece=ece,eee=eee,fed=fed,mba=mba,mca=mca,ce=ce,me=me,ae=ae,others=others)
        x = mysql.connection.cursor()
        x.execute('select * from user1 where id=%s',(id,))
        y = x.fetchall()
        x.close()
        gc.collect()
        count = 0
        for i in y:
            count = count + 1
        if count > 0:
            x = mysql.connection.cursor()
            x.execute('delete from user1 where id = %s',(id,))
            x.connection.commit()
            x.close()
            flash('Deleted successfullt..')
            gc.collect()
            return render_template('viewfaculty.html',username=username,script=script,cse=cse,it=it,ece=ece,eee=eee,fed=fed,mba=mba,mca=mca,ce=ce,me=me,ae=ae,others=others)
        else:
            flash('Faculty account for entered id is not available...')
            return render_template('viewfaculty.html',username=username,script=script,cse=cse,it=it,ece=ece,eee=eee,fed=fed,mba=mba,mca=mca,ce=ce,me=me,ae=ae,others=others)
    return render_template('viewfaculty.html',username=username,script=script,cse=cse,it=it,ece=ece,eee=eee,fed=fed,mba=mba,mca=mca,ce=ce,me=me,ae=ae,others=others)


    return render_template('casformset.html',username=username)

@app.route('/casreset', methods=['GET','POST'])
def casreset():
    username = session.get('username',None)
    x = mysql.connection.cursor()
    flag = x.execute('select * from casmarks')
    x.close()
    gc.collect()
    if flag == False:
        flash('No data entered using CAS Scores SET ... !!!')
        return render_template('cas.html',username=username)
    else:
        
        if request.method == 'POST':
            ugc = request.form['indexugc']
            scopus = request.form['indexscopus']
            sci = request.form['indexsci']
            oth = request.form['indexother']
            iit = request.form['pointiit']
            iiit = request.form['pointiiit']
            nit = request.form['pointnit']
            univ = request.form['pointuniv']
            duniv = request.form['pointduniv']
            puniv = request.form['pointpuniv']
            clg = request.form['pointother']
            x = mysql.connection.cursor()
            x.execute('delete from casmarks where cid = 0')
            flag = x.execute('select * from casmarks')
            x.close()
            gc.collect()
            if flag == False:
                x = mysql.connection.cursor()
                s = 'insert into casmarks(ugc,scopus,sci,others,iit,iiit,nit,univ,duniv,puniv,college) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
                a = (ugc,scopus,sci,oth,iit,iiit,nit,univ,duniv,puniv,clg,)
                x.execute(s,a)
                x.connection.commit()
                x.close()
                gc.collect()
                flash('CAS Scores Updated successfully ... !!!')
                return render_template('cas.html',username=username)
    return render_template('casreset.html',username=username)


extensions = set(['txt','pdf','doc','docx','jpg','jpeg','png','xlsx'])
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in extensions


def convertToBinaryData(filename):
    #Convert digital data to binary format
    with open(filename, 'rb') as file:
        binaryData = file.read()
    return binaryData

def insertBLOB(username,filename,filedescription,file1):
    x = mysql.connection.cursor()
    sql_insert_blob_query = """ INSERT INTO `filestore`(`id`, `filename`, `filedescription`, `file`) VALUES (%s,%s,%s,%s)"""
    file = convertToBinaryData(file1)
    # Convert data into tuple format
    insert_blob_tuple = (username, filename, filedescription, file)
    result  = cursor.execute(sql_insert_blob_query, insert_blob_tuple)
    x.connection.commit()
    x.close()
    gc.collect

'''

@app.route('/upload',methods=['GET','POST'])
def upload():
    username = session.get('username',None)
    if request.method == 'POST':
        filename = request.form['title']
        filedescription = request.form['description']
        file = request.files['file']
        if file.filename == '':
            flash('To upload, first select a proper file ... !!!')
            return render_template('upload.html',username=username)
        if file and allowed_file(file.filename):
            with open(file.filename, 'rb') as file:
                binaryData = file.read()
            x = mysql.connection.cursor()
            x.execute('insert into filestore(id,filename,filedescription,file) values (%s,%s,%s,%s)'(username,filename,filedescription,binaryData))
            x.connection.commit()
            x.close()
            gc.collect()
            flash('Image uploaded successfully... !!!')
            return render_template('upload.html',username=username)
    else:
        return render_template('upload.html',username=username)


'''
@app.route('/upload',methods=['GET','POST'])
def upload():
    username = session.get('username',None)
    folder = "C:\\Users\\Sri Syam\\myproj\\uploads"
    app.config['UPLOAD_FOLDER'] = folder
    if request.method == 'POST':
        filename = request.form['title']
        fname = request.form['title']
        filedescription = request.form['description']
        file = request.files['file']
        if file.filename == '':
            flash('To upload, first select a proper file ... !!!')
            return render_template('upload.html',username=username)
        if filedescription == '':
            flash('Enter valid file description ... !!!')
            return render_template('upload.html',username=username)
        if file and allowed_file(file.filename):
            if os.path.isdir(os.path.join(app.config['UPLOAD_FOLDER'],username)):
                filename = secure_filename(file.filename)
                '''
                x = mysql.connection.cursor()
                x.execute('select filename from filestore where id=%s',(username,))
                y = x.fetchall()
                x.connection.close()
                gc.collect()
                if filename in y:
                    flash(' The uploading already exists ...')
                    return render_template('upload.html',username=username)
                '''
                file.save(os.path.join(os.path.join(app.config['UPLOAD_FOLDER'],username),filename))
                dirname = os.path.join(os.path.join(app.config['UPLOAD_FOLDER'],username),filename)
                flash(' File uploaded successfully .... ')
                x = mysql.connection.cursor()
                s1 = 'insert into filestore(id,fname,filename,filedescription,filepath) values (%s,%s,%s,%s,%s)'
                args = (username,fname,filename,filedescription,dirname)
                x.execute(s1,args)
                x.connection.commit()
                x.close()
                gc.collect()
                return render_template('upload.html',username=username)
            else:
                os.mkdir(os.path.join(app.config['UPLOAD_FOLDER'],username))
                filename = secure_filename(file.filename)
                '''
                x = mysql.connection.cursor()
                x.execute('select filename from filestore where id=%s',(username,))
                y = x.fetchall()
                x.connection.close()
                gc.collect()
                if filename in y:
                    flash(' The uploading already exists ...')
                    return render_template('upload.html',username=username)
                '''
                file.save(os.path.join(os.path.join(app.config['UPLOAD_FOLDER'],username),filename))
                dirname = os.path.join(os.path.join(app.config['UPLOAD_FOLDER'],username),filename)
                x = mysql.connection.cursor()
                s1 = 'insert into filestore(id,fname,filename,filedescription,filepath) values (%s,%s,%s,%s,%s)'
                args = (username,fname,filename,filedescription,dirname)
                x.execute(s1,args)
                x.connection.commit()
                x.close()
                gc.collect()
                flash(' File uploaded successfully .... ')
                return render_template('upload.html',username=username)
        else:
            flash("Error in allowed file extensions")
            return render_template('upload.html',username=username)
    return render_template('upload.html',username=username)


def write_file(data, filename):
    # Convert binary data to proper format and write it on Hard Disk
    with open(filename, 'wb') as file:
        file.write(data)


@app.route('/viewupload', methods=['GET','POST'])
def viewupload():
    username = session.get('username',None)
    x = mysql.connection.cursor()
    x.execute('select * from filestore where id=%s',(username,))
    y = x.fetchall()
    rows = x.rowcount
    x.close()
    gc.collect()
    value = ''
    type = ''
    getvalue=''
    if request.method == 'POST':
        for i in range(0,rows):
            if request.form['submit'] == 'view'+str(i):
                value = i
                type = 'view'
                getvalue = request.form['t'+str(i)]
                return render_template('viewpage.html',username=username,rows=value,type=type,getvalue=getvalue)
            elif request.form['submit'] == 'download'+str(i):
                value = i
                type = 'download'
                getvalue = request.form['t'+str(i)]
                x = mysql.connection.cursor()
                x.execute('select filepath from filestore where fname=%s',(getvalue,))
                b = x.fetchall()
                x.close()
                gc.collect()
                str1 = str(b[0])
                str2= str1[2:-3]
                os.path.normpath(str2)
                return send_file(str2,as_attachment=True)
            elif request.form['submit'] == 'delete'+str(i):
                value = i
                type = 'delete'
                getvalue = request.form['t'+str(i)]
                x = mysql.connection.cursor()
                x.execute('select filepath from filestore where fname=%s',(getvalue,))
                b = x.fetchall()
                x.execute('delete from filestore where fname=%s',(getvalue,))
                x.connection.commit()
                x.close()
                gc.collect()
                str1 = str(b[0])
                str2= str1[2:-3]
                os.path.normpath(str2)
                os.remove(str2)
                x = mysql.connection.cursor()
                x.execute('select * from filestore where id=%s',(username,))
                y = x.fetchall()
                rows = x.rowcount
                x.close()
                gc.collect()
                flash('Record deleted successfully .... ')
                return render_template('viewupload.html',username=username,y=y)
    return render_template('viewupload.html',username=username,y=y)

@app.route('/publications',methods=['GET','POST'])
def publications():
    username=session.get('username',None)
    if request.method == 'POST':
        select = request.form['pubsel']
        if select == 'none':
            flash('Select proper option ... !!!!')
            return render_template('publications.html',username=username)
        elif select == 'conference':
            return render_template('conference.html',username=username)
        elif select == 'journal':
            return render_template('journal.html',username=username)
        elif select == 'viewpublish':
            return render_template('viewpublish.html',username=username)
        elif select == 'delcon':
            x = mysql.connection.cursor()
            x.execute('select * from conference where id=%s',(username,))
            y = x.fetchall()
            clen = len(y)
            x.close()
            gc.collect()
            return render_template('delcon.html',username=username,y=y,clen=clen)
        elif select == 'delpub':
            x = mysql.connection.cursor()
            x.execute('select * from journal where id=%s',(username,))
            b = x.fetchall()
            jlen = len(b)
            x.close()
            gc.collect()
            return render_template('delpub.html',username=username,b=b,jlen=jlen)
    return render_template('publications.html',username=username)

@app.route('/conference',methods=['GET','POST'])
def conference():
    username = session.get('username',None)
    if request.method == 'POST':
        title = request.form['title']
        names = request.form['names']
        barea = request.form['barea']
        ctitle = request.form['ctitle']
        ind = request.form['index']
        score=0
        if ind =='ugc':
            x = mysql.connection.cursor()
            x.execute('select ugc from casmarks')
            score = x.fetchall()
            x.close()
            gc.collect()
        elif ind =='scopus':
            x = mysql.connection.cursor()
            x.execute('select scopus from casmarks')
            score = x.fetchall()
            x.close()
            gc.collect()
        elif ind =='sci':
            x = mysql.connection.cursor()
            x.execute('select sci from casmarks')
            score = x.fetchall()
            x.close()
            gc.collect()
        else :
            x = mysql.connection.cursor()
            x.execute('select others from casmarks')
            score = x.fetchall()
            x.close()
            gc.collect()
        cdate = request.form['cdate']
        venue = request.form['venue']
        jname = request.form['jname']
        issn = request.form['issn']
        volume = request.form['volume']
        issue = request.form['issue']
        page = request.form['page']
        if title == '':
            flash('Enter Title of the research paper')
            return render_template('conference.html',username=username)
        elif names == '':
             flash('Enter atleast an author name')
             return render_template('conference.html',username=username)
        elif ctitle == '':
            flash('Enter conferenece title')
            return render_template('conference.html',username=username)
        elif cdate == '':
            flash('Enter conferenece date')
            return render_template('conference.html',username=username)
        else:
            cdate=datetime.strptime(cdate,"%Y-%m-%d")
            s = "insert into conference(id,title,names,barea,ctitle,ind,cdate,venue,jname,issn,volume,issue,page,score) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
            a = (username,title,names,barea,ctitle,ind,(cdate.strftime('%Y-%m-%d')),venue,jname,issn,volume,issue,page,score)
            x = mysql.connection.cursor()
            x.execute(s,a)
            x.connection.commit()
            x.close()
            gc.collect()
            flash('publication details entered successfully ... !!!')
            return render_template('publications.html',username=username)
    return render_template('conference.html',username=username)

@app.route('/journal',methods=['GET','POST'])
def journal():
    username = session.get('username',None)
    if request.method == 'POST':
        title = request.form['title']
        names = request.form['names']
        barea = request.form['barea']
        ind = request.form['index']
        score=0
        if ind =='ugc':
            x = mysql.connection.cursor()
            x.execute('select ugc from casmarks')
            score = x.fetchall()
            x.close()
            gc.collect()
        elif ind =='scopus':
            x = mysql.connection.cursor()
            x.execute('select scopus from casmarks')
            score = x.fetchall()
            x.close()
            gc.collect()
        elif ind =='sci':
            x = mysql.connection.cursor()
            x.execute('select sci from casmarks')
            score = x.fetchall()
            x.close()
            gc.collect()
        else :
            x = mysql.connection.cursor()
            x.execute('select others from casmarks')
            score = x.fetchall()
            x.close()
            gc.collect()
        cdate = request.form['cdate']
        jname = request.form['jname']
        issn = request.form['issn']
        volume = request.form['volume']
        issue = request.form['issue']
        page = request.form['page']
        if title == '':
            flash('Enter Title of the research paper')
            return render_template('journal.html',username=username)
        elif names == '':
             flash('Enter atleast an author name')
             return render_template('journal.html',username=username)
        elif jname == '':
            flash('Enter journal title')
            return render_template('journal.html',username=username)
        elif cdate == '':
            flash('Enter publication date')
            return render_template('journal.html',username=username)
        else:
            cdate=datetime.strptime(cdate,"%Y-%m-%d")
            s = "insert into journal(id,title,names,barea,ind,cdate,jname,issn,volume,issue,page,score) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
            a = (username,title,names,barea,ind,(cdate.strftime('%Y-%m-%d')),jname,issn,volume,issue,page,score)
            x = mysql.connection.cursor()
            x.execute(s,a)
            x.connection.commit()
            x.close()
            gc.collect()
            flash('publication details entered successfully ... !!!')
            return render_template('publications.html',username=username)
    return render_template('journal.html',username=username)

@app.route('/viewpublish',methods=['GET','POST'])
def viewpublish():
    username = session.get('username',None)
    if request.method == 'POST':
        fromdate = request.form['fromdate']
        todate = request.form['todate']
        if fromdate > todate:
            flash('From-Date must be less than To-Date')
            return render_template('viewpublish.html',username=username)
        else:
            x = mysql.connection.cursor()
            x.execute("select * from conference where id=%s and cdate>=%s and cdate<=%s",(username,fromdate,todate,))
            y = x.fetchall()
            clen = len(y)
            x.close()
            gc.collect()
            a = mysql.connection.cursor()
            a.execute("select * from journal where id=%s and cdate>=%s and cdate<=%s",(username,fromdate,todate,))
            b = a.fetchall()
            jlen = len(b)
            a.close()
            gc.collect()
            return render_template('viewpublishcontent.html',username=username,y=y,b=b,clen=clen,jlen=jlen)
        return render_template('viewpublish.html',username=username)
    return render_template('viewpublish.html',username=username)

@app.route('/delcon',methods=['GET','POST'])
def delcon():
    username = session.get('username',None)
    x = mysql.connection.cursor()
    x.execute('select * from conference where id=%s',(username,))
    y = x.fetchall()
    clen = len(y)
    x.close()
    gc.collect()
    if request.method == 'POST':
        refid = request.form['refid']
        if refid == '':
            flash('Enter a valid refid ... !!!')
            return render_template('delcon.html',y=y,clen=clen,username=username)
        x = mysql.connection.cursor()
        dd=x.execute('delete from conference where refid = %s and id = %s',(refid,username,))
        x.connection.commit()
        x.close()
        gc.collect()
        if dd == False:
            flash('Enter Valid RefID ... !!!')
        else:
            flash('Conference deleted successfully ... !!! Number deleted is ')
        x = mysql.connection.cursor()
        x.execute('select * from conference where id=%s',(username,))
        y = x.fetchall()
        clen = len(y)
        x.close()
        gc.collect()
        return render_template('delcon.html',y=y,clen=clen,username=username)
    return render_template('delcon.html',username=username,y=y,clen=clen)


@app.route('/delpub',methods=['GET','POST'])
def delpub():
    username = session.get('username',None)
    x = mysql.connection.cursor()
    x.execute('select * from journal where id=%s',(username,))
    b = x.fetchall()
    jlen = len(b)
    x.close()
    gc.collect()
    if request.method == 'POST':
        refid = request.form['refid']
        if refid == '':
            flash('Enter a valid refid ... !!!')
            return render_template('delpub.html',b=b,jlen=jlen,username=username)
        x = mysql.connection.cursor()
        dd=x.execute('delete from journal where jrefid = %s and id = %s',(refid,username,))
        x.connection.commit()
        x.close()
        gc.collect()
        if dd == False:
            flash('Enter Valid RefID ... !!!')
        else:
            flash('Publicaion deleted successfully ... !!!')
        x = mysql.connection.cursor()
        x.execute('select * from journal where id=%s',(username,))
        b = x.fetchall()
        jlen = len(b)
        x.close()
        gc.collect()
        return render_template('delpub.html',b=b,jlen=jlen,username=username)
    return render_template('delpub.html',username=username,b=b,jlen=jlen)

@app.route('/fdp',methods=['GET','POST'])
def fdp():
    username = session.get('username',None)
    if request.method == 'POST':
        sel = request.form['fdpsel']
        if sel == 'none':
            flash('Please select a service to proceed ... !!!')
            return render_template("fdp.html",username=username)
        elif sel == 'addfdp':
            return render_template('addfdp.html',username=username)
        elif sel == 'viewfdp':
            return render_template('viewfdp.html',username=username)
        elif sel == 'delfdp':
            x = mysql.connection.cursor()
            x.execute('select * from fdp where id=%s',(username,))
            y = x.fetchall()
            flen = len(y)
            x.close()
            gc.collect()
            return render_template('delfdp.html',y=y,flen=flen,username=username)
    return render_template('fdp.html',username=username)


@app.route('/addfdp',methods=['GET','POST'])
def addfdp():
    username = session.get('username',None)
    if request.method == 'POST':
        fdptitle = request.form['fdptitle']
        fromdate = request.form['fromdate']
        fromdate = datetime.strptime(fromdate,"%Y-%m-%d")
        todate = request.form['todate']
        todate = datetime.strptime(todate,"%Y-%m-%d")
        venuetype = request.form['venuetype']
        venue = request.form['venue']
        if fdptitle == '':
            flash('Enter a valid title')
            return render_template('addfdp.html',username=username)
        if fromdate > todate:
            flash('From date muse be greater than to date ... !!!')
            return render_template('addfdp.html',username=username)
        if venue == '':
            flash('Enter a valid venue')
            return render_template('addfdp.html',username=username)
        score=0
        if venuetype =='iit':
            x = mysql.connection.cursor()
            x.execute('select iit from casmarks')
            score = x.fetchall()
            x.close()
            gc.collect()
        elif venuetype =='iiit':
            x = mysql.connection.cursor()
            x.execute('select iiit from casmarks')
            score = x.fetchall()
            x.close()
            gc.collect()
        elif venuetype =='nit':
            x = mysql.connection.cursor()
            x.execute('select nit from casmarks')
            score = x.fetchall()
            x.close()
            gc.collect()
        elif venuetype =='univ':
            x = mysql.connection.cursor()
            x.execute('select univ from casmarks')
            score = x.fetchall()
            x.close()
            gc.collect()
        elif venuetype =='duniv':
            x = mysql.connection.cursor()
            x.execute('select duniv from casmarks')
            score = x.fetchall()
            x.close()
            gc.collect()
        elif venuetype =='puniv':
            x = mysql.connection.cursor()
            x.execute('select puniv from casmarks')
            score = x.fetchall()
            x.close()
            gc.collect()
        elif venuetype =='college':
            x = mysql.connection.cursor()
            x.execute('select college from casmarks')
            score = x.fetchall()
            x.close()
            gc.collect()
        daynum = todate-fromdate
        daycount = daynum.days
        daycount = daycount + 1
        x = mysql.connection.cursor()
        x.execute('insert into fdp(id,fdptitle,fromdate,todate,venuetype,venue,days,score) values(%s,%s,%s,%s,%s,%s,%s,%s)',(username,fdptitle,fromdate.strftime('%Y-%m-%d'),todate.strftime('%Y-%m-%d'),venuetype,venue,int(daycount),score))
        x.connection.commit()
        x.close()
        gc.collect()
        flash('FDP entered successfully ... !!!')
        return render_template('fdp.html',username=username)
    return render_template('addfdp.html',username=username)


@app.route('/viewfdp',methods=['GET','POST'])
def viewfdp():
    username = session.get('username',None)
    if request.method == 'POST':
        fromdate = request.form['fromdate']
        todate = request.form['todate']
        if fromdate > todate:
            flash('From-Date must be less than To-Date')
            return render_template('viewfdp.html',username=username)
        else:
            x = mysql.connection.cursor()
            x.execute("select * from fdp where id=%s and fromdate>=%s and fromdate<=%s",(username,fromdate,todate,))
            y = x.fetchall()
            flen = len(y)
            x.close()
            gc.collect()
            return render_template('viewfdpcontent.html',username=username,y=y,flen=flen)
        return render_template('viewfdp.html',username=username)
    return render_template('viewfdp.html',username=username)

@app.route('/delfdp',methods=['GET','POST'])
def delfdp():
    username = session.get('username',None)
    x = mysql.connection.cursor()
    x.execute('select * from fdp where id=%s',(username,))
    y = x.fetchall()
    flen = len(y)
    x.close()
    gc.collect()
    if request.method == 'POST':
        refid = request.form['refid']
        x = mysql.connection.cursor()
        flag = x.execute('delete from fdp where id=%s and refid=%s',(username,refid,))
        x.connection.commit()
        x.close()
        gc.collect()
        if flag == False:
            flash('Enter valid RefID to delete ... !!!')
            return render_template('delfdp.html',y=y,flen=flen,username=username)
        else:
            x = mysql.connection.cursor()
            x.execute('select * from fdp where id=%s',(username,))
            y = x.fetchall()
            flen = len(y)
            x.close()
            gc.collect()
            flash('FDP deleted successfully ... !!!')
            return render_template('delfdp.html',y=y,flen=flen,username=username)
    return render_template('delfdp.html',y=y,flen=flen,username=username)

@app.route('/hodviewfac',methods=['GET','POST'])
def hodviewfac():
    username = session.get('username',None)
    l = []
    fcount = 0
    faculty = []
    slen = len(username)
    s2 = username[3:]
    x = mysql.connection.cursor()
    flag = x.execute('select id from user1')
    y = x.fetchall()
    flen  = len(y)
    if flag == False:
        flash('No faculy registered ... !!!')
        return render_template('hodviewfac.html',username=username)
    else:
        count = 0
        for i in y:
            l.append(i)
            count = count + 1
        if count > 0:
            l.sort()
            for i in l:
                j=''
                j =''.join(i)
                if j.startswith(s2,0,len(s2)):
                    fcount = fcount + 1
                    faculty.append(j)
        return render_template('hodviewfac.html',username=username,fcount=fcount,faculty=faculty)
    return render_template('hodviewfac.html',username=username)

@app.route('/hodviewcpfdate', methods=['GET','POST'])
def hodviewcpfdate():
    username = session.get('username',None)
    if request.method == 'POST':
        fromdate = request.form['fromdate']
        todate = request.form['todate']
        if fromdate == '' or todate == '':
            flash('From date and to date must not empty ... !!!')
            return render_template('hodviewcpfdate.html',username=username)
        if fromdate > todate:
            flash('Enter valid from date and to dates ... !!!')
            return render_template('hodviewcpfdate.html',username=username)
        l = []
        conf = 0
        jour = 0
        fdp = 0
        fcount = 0
        faculty = []
        cdict = dict()
        cdictu = dict()
        cdicts = dict()
        cdictsci = dict()
        cdicto = dict()
        jdict = dict()
        jdictu = dict()
        jdicts = dict()
        jdictsci = dict()
        jdicto = dict()
        fdict = dict()
        fdicti = dict()
        fdictii = dict()
        fdictn = dict()
        fdictu = dict()
        fdictd = dict()
        fdictp = dict()
        fdictc = dict()
        slen = len(username)
        s2 = username[3:]
        UGC = 'ugc'
        x = mysql.connection.cursor()
        flag = x.execute('select id from user1')
        y = x.fetchall()
        x.close()
        gc.collect()
        flen  = len(y)
        if flag == False:
            flash('No faculy registered ... !!!')
            return render_template('hodviewcpf.html',username=username)
        else:
            count = 0
            for i in y:
                l.append(i)
                count = count + 1
            if count > 0:
                l.sort()
                for i in l:
                    j=''
                    j =''.join(i)
                    if j.startswith(s2,0,len(s2)):
                        fcount = fcount + 1
                        faculty.append(j)
                for i in range(0,fcount):
                    cdict[faculty[i]] = 0
                    cdictu[faculty[i]] = 0
                    cdicts[faculty[i]] = 0
                    cdictsci[faculty[i]] = 0
                for i in range(0,fcount):
                    x = mysql.connection.cursor()
                    x.execute('select * from conference where id =%s and cdate>=%s and cdate <= %s',(faculty[i],fromdate,todate,))
                    y = x.fetchall()
                    cdict[faculty[i]] = len(y)
                    x.close()
                    gc.collect()
                    #cdictu,cdicts,cdictsci,sdicto
                    x = mysql.connection.cursor()
                    x.execute('select * from conference where id=%s and ind=%s and cdate>=%s and cdate<=%s ',(faculty[i],UGC,fromdate,todate,))
                    y = x.fetchall()
                    cdictu[faculty[i]] = len(y)
                    x.close()
                    gc.collect()
                    x = mysql.connection.cursor()
                    x.execute('select * from conference where id=%s and ind = %s and cdate>=%s and cdate<=%s',(faculty[i],'scopus',fromdate,todate,))
                    y = x.fetchall()
                    cdicts[faculty[i]] = len(y)
                    x.close()
                    gc.collect()
                    x = mysql.connection.cursor()
                    x.execute('select * from conference where id=%s and ind = %s and cdate>=%s and cdate<=%s',(faculty[i],'sci',fromdate,todate,))
                    y = x.fetchall()
                    cdictsci[faculty[i]] = len(y)
                    x.close()
                    gc.collect()
                    x = mysql.connection.cursor()
                    x.execute('select * from conference where id=%s and cdate>=%s and cdate<=%s and ind =%s',(faculty[i],fromdate,todate,'other',))
                    y = x.fetchall()
                    cdicto[faculty[i]] = len(y)
                    x.close()
                    gc.collect()
                for i in range(0,fcount):
                    x = mysql.connection.cursor()
                    x.execute('select * from journal where id =%s and cdate>=%s and cdate <= %s',(faculty[i],fromdate,todate,))
                    y = x.fetchall()
                    jdict[faculty[i]] = len(y)
                    x.close()
                    gc.collect()
                for i in range(0,fcount):
                    x = mysql.connection.cursor()
                    x.execute('select * from fdp where id =%s and fromdate>=%s and fromdate <= %s',(faculty[i],fromdate,todate,))
                    y = x.fetchall()
                    fdict[faculty[i]] = len(y)
                    x.close()
                    gc.collect()
            return render_template('hodviewcpf.html',username=username,fcount=fcount,faculty=faculty,cdict=cdict,jdict=jdict,fdict=fdict,cdictu=cdictu,cdicts=cdicts,cdictsci=cdicts,cdicto=cdicto)
        return render_template('hodviewcpf.html',username=username)
    return render_template('hodviewcpfdate.html',username=username)

@app.route('/hodviewcpf',methods=['GET','POST'])
def hodviewcpf():
    username = session.get('username',None)
    l = []
    conf = 0
    jour = 0
    fdp = 0
    fcount = 0
    faculty = []
    cdict = dict()
    jdict = dict()
    fdict = dict()
    slen = len(username)
    s2 = username[3:]
    x = mysql.connection.cursor()
    flag = x.execute('select id from user1')
    y = x.fetchall()
    x.close()
    gc.collect()
    flen  = len(y)
    if flag == False:
        flash('No faculy registered ... !!!')
        return render_template('hodviewcpf.html',username=username)
    else:
        count = 0
        for i in y:
            l.append(i)
            count = count + 1
        if count > 0:
            l.sort()
            for i in l:
                j=''
                j =''.join(i)
                if j.startswith(s2,0,len(s2)):
                    fcount = fcount + 1
                    faculty.append(j)
            for i in range(0,fcount):
                cdict[faculty[i]] =0
            for i in range(0,fcount):
                x = mysql.connection.cursor()
                x.execute('select * from conference where id =%s',(faculty[i],))
                y = x.fetchall()
                cdict[faculty[i]] = len(y)
                x.close()
                gc.collect()
            for i in range(0,fcount):
                x = mysql.connection.cursor()
                x.execute('select * from journal where id =%s',(faculty[i],))
                y = x.fetchall()
                jdict[faculty[i]] = len(y)
                x.close()
                gc.collect()
            for i in range(0,fcount):
                x = mysql.connection.cursor()
                x.execute('select * from fdp where id =%s',(faculty[i],))
                y = x.fetchall()
                fdict[faculty[i]] = len(y)
                x.close()
                gc.collect()
        return render_template('hodviewcpf.html',username=username,fcount=fcount,faculty=faculty,cdict=cdict,jdict=jdict,fdict=fdict)
    return render_template('hodviewcpf.html',username=username)

if __name__ == '__main__':
    app.run(debug=True)
